#!/usr/bin/env python3
"""
导出 Supabase 所有 public 表为 Excel 文件到 public/downloads/ 目录
用法: python scripts/export_all_tables.py [--output-dir public/downloads]
"""
import os, sys, json, requests
from datetime import datetime

# ===== 配置 =====
SUPABASE_URL = os.environ.get('SUPABASE_URL') or os.environ.get('VITE_SUPABASE_URL') or 'https://tqhtegazxykkqfcpejky.supabase.co'
SUPABASE_KEY = os.environ.get('SUPABASE_ANON_KEY') or os.environ.get('VITE_SUPABASE_ANON_KEY') or 'sb_publishable_iFtMcvav774gqF28gGYQVw_QMmuS-z3'

# 判断是否 CI 环境
IS_CI = os.environ.get('CI') == 'true' or os.environ.get('GITHUB_ACTIONS') == 'true'
if IS_CI:
    import subprocess
    result = subprocess.run(['pip', 'install', 'openpyxl', 'requests'], capture_output=True, text=True)
    print(f"[CI] pip install openpyxl requests: {result.returncode}")
    if result.returncode != 0:
        print(f"[CI] pip stderr: {result.stderr}")
        sys.exit(1)
else:
    try:
        import openpyxl
    except ImportError:
        print("请先安装 openpyxl: pip install openpyxl requests")
        sys.exit(1)

import openpyxl
from openpyxl.utils import get_column_letter

# 输出目录
OUTPUT_DIR = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), 'public', 'downloads')
if '--output-dir' in sys.argv:
    idx = sys.argv.index('--output-dir')
    OUTPUT_DIR = sys.argv[idx + 1]
os.makedirs(OUTPUT_DIR, exist_ok=True)

HEADERS = {
    'apikey': SUPABASE_KEY,
    'Authorization': f'Bearer {SUPABASE_KEY}',
    'Content-Type': 'application/json',
}

# ===== 表定义 =====
TABLES = {
    'fund_combined': {
        'name': '基金综合数据表',
        'desc': '基金分类(t0/t1)、公司/规模/费率、收益(ytd/r1y/r3y/r5y)、风险指标(dd1y/sr1y)、持有人数、评分(k_all/score_grade/k0w~k10) — 所有数据核心合并表，19+ 周期评分全覆盖',
        'source': '天天基金 FundGuideapi（收益率/分类）+ pingzhongdata（回撤/夏普/风险评级）+ rankhandler（货币基金收益）+ fundf10（公司/规模/费率）',
        'update': '每日通过 GitHub Actions CI 自动更新（北京时间 21:30）',
        'scoring': True,
    },
    'fund_scores': {
        'name': '基金评分表（完整版）',
        'desc': '每日更新的核心数据表：基金代码/名称/基金经理/管理人/分类/规模/费率 → 阶段收益(ytd/r0w~r10y/return_all) → 阶段回撤(dd1y~dd5y) → 阶段夏普(sr1y~sr5y) → 基金评分(k0w~k_all/score_grade)。按以上顺序排列。',
        'source': 'FundGuideapi（收益率/分类）+ pingzhongdata（回撤/夏普/基金经理）+ fund_combined（公司/规模/费率）+ rankhandler（货币基金/成立以来收益）',
        'update': '每日通过 GitHub Actions CI 自动更新（北京时间 21:30）',
        'scoring': True,
    },
    'fund_indices': {
        'name': '基金指数表（万得 Wind）',
        'desc': '万得(Wind)基金指数：代码/名称/分类/类型 + 基本信息(发布日期/成分数量/加权方式/收益方式) + 市场表现(近1周~成立以来收益率) + 历年表现(年度收益) + 估值分析(总市值/流通市值/市盈率/净利率/股息率/Beta/波动率/换手率)',
        'source': 'windindices.com（万得基金指数），通过 scripts/fetch_fund_indices.py 浏览器会话抓取（方案A）',
        'update': '通过 scripts/fetch_fund_indices.py 抓取更新',
        'scoring': False,
    },
    'fund_scores_test': {
        'name': '基金评分测试表',
        'desc': 'fund_scores 的测试副本，结构与 fund_scores 完全一致。新抓取数据先写入此表验证无误后，再通过 staging 管道导入 fund_scores 生产环境。',
        'source': 'CI 抓取流程写入（验证用）',
        'update': '每次抓取数据时先写入此表',
        'scoring': True,
    },
    'index_eva': {
        'name': '行业估值表（生产）',
        'desc': '蛋卷指数估值数据：指数代码/名称/类型(宽基/策略/行业主题)/PE/PB/股息率/ROE/PE历史分位/PB历史分位/估值评级。网页指标信号页行业估值与生产表一致。',
        'source': '蛋卷基金 danjuanfunds.com（index_eva/dj 接口）',
        'update': '每日通过 GitHub Actions CI 自动更新（北京时间 21:30），先增量写入 index_eva_test 验证再同步生产',
        'scoring': False,
    },
    'index_eva_test': {
        'name': '行业估值表（测试）',
        'desc': 'index_eva 的测试副本，结构与生产表一致。每次抓取数据先以增量断点续传方式写入此表验证无误后再同步到 index_eva 生产环境。',
        'source': 'CI 抓取流程写入（验证用）',
        'update': '每次抓取数据时先写入此表',
        'scoring': False,
    },
    'fund_quarterly_scores': {
        'name': '季度评分表',
        'desc': '基于季报数据的各时间窗口评分（score_3m/6m/1y/2y/3y/5y/7y/10y），含原始 quarterly_data JSON',
        'source': 'pingzhongdata 每日净值 → 季度收益/回撤/夏普计算 → 全市场排名 → 多周期均值评分',
        'update': '每日通过 GitHub Actions CI 自动更新（北京时间 21:30）',
        'scoring': False,
    },
    'fund_scores_meta': {
        'name': '评分元数据表',
        'desc': '评分更新时间、基金总数、有评分数、净值日期等元信息',
        'source': 'Supabase 内部自动记录',
        'update': '每次评分计算完成后自动更新',
        'scoring': False,
    },
    'config': {
        'name': '配置表',
        'desc': '全站配置项（键值对，含 meta、tsq 时间戳）',
        'source': '手动维护',
        'update': '按需手动更新',
        'scoring': False,
    },
    'macro_history': {
        'name': '宏观历史数据表',
        'desc': '中国10年国债(cn10y)、美国10年国债(us10y)、Shibor、CPI、M2 的历史数据，覆盖 1996-至今',
        'source': 'akshare 开源 Python 库（自动采集公开宏观数据）',
        'update': '每日通过 GitHub Actions CI 自动更新（北京时间 21:30）',
        'scoring': False,
    },
    'index_pe_history': {
        'name': '指数PE历史表',
        'desc': '沪深300等指数的 PE/PB 历史估值数据',
        'source': '腾讯行情 qt.gtimg.cn + 蛋卷基金 danjuanfunds.com',
        'update': '每日通过 GitHub Actions CI 自动更新（北京时间 21:30）',
        'scoring': False,
    },
    'site_stats': {
        'name': '站点统计表',
        'desc': '网站访问量等统计指标',
        'source': 'EdgeOne Pages 边缘函数自动记录',
        'update': '实时更新',
        'scoring': False,
    },
    'tougu_products': {
        'name': '投顾产品表',
        'desc': '天天基金/华宝/盈米/新浪仓石四来源的基金投顾产品，含收益率、最大回撤、标签分类',
        'source': '天天基金投顾页面 + 华宝/盈米/新浪仓石官方数据',
        'update': '每日通过 GitHub Actions CI 自动更新（北京时间 21:30）',
        'scoring': False,
    },
    'user_portfolios': {
        'name': '用户组合表',
        'desc': '用户自建基金组合数据（portfolio_data JSON），关联用户 ID',
        'source': '用户通过 大厨先生 网站自行创建',
        'update': '用户操作时实时更新',
        'scoring': False,
    },
    'user_profiles': {
        'name': '用户档案表',
        'desc': '用户注册信息、登录次数、最后登录时间',
        'source': '用户注册时填写',
        'update': '用户操作时实时更新',
        'scoring': False,
    },
    'fund_tags': {
        'name': '热门标签表',
        'desc': '热门基金标签（行业 hy + 概念 gn），含标签名/类型/近1年板块收益/排序',
        'source': '东财 ZTJJ GetBKListByBKTypeNew 接口（行业/概念标签清单）',
        'update': '通过 scripts/fetch_tag_funds_v2.py 抓取更新',
        'scoring': False,
    },
    'fund_tag_funds': {
        'name': '标签-基金映射表',
        'desc': '每个热门标签关联的基金列表（基金代码/名称/类型/近1年收益/日涨跌/排序）',
        'source': '东财 ZTJJ GetBKRelTopicFundNew 接口（标签关联基金）',
        'update': '通过 scripts/fetch_tag_funds_v2.py 抓取更新',
        'scoring': False,
    },
    'fund_tag_perf': {
        'name': '主题板块涨跌表',
        'desc': '154个热门行业/概念板块的板块级涨跌幅：日涨跌(D)/近1周(W)/近1月(M)/近3月(Q)/近1年(Y)/今年来(SY) + 各周期排名 + 板块总数。热门基金「实时/近1周/近1月/近3月/近1年/今年来」排序的数据基础',
        'source': '东财 ZTJJ GetBKDetailInfoNew 接口（板块级真实涨跌，非基金个体收益均值）',
        'update': '通过 scripts/sync_tag_performance.py 抓取更新，每日 GitHub Actions CI 自动刷新（北京时间 21:30）',
        'scoring': False,
    },
    'ai_pk_models': {
        'name': 'AI大PK 模型表',
        'desc': 'AI 大PK 参赛模型信息（模型名/厂商/头像/描述/状态等）',
        'source': '大厨先生 内部配置 + 各大模型 API',
        'update': '按需更新',
        'scoring': False,
    },
    'ai_pk_picks': {
        'name': 'AI大PK 选基表',
        'desc': '各 AI 模型每期选出的基金及权重（基于 fund_scores 真实数据）',
        'source': '各大模型 API（DeepSeek/豆包/智谱等）基于 fund_scores 选基',
        'update': '按调仓周期更新',
        'scoring': False,
    },
    'factor_scores': {
        'name': '风格因子评分表（生产）',
        'desc': '股票/债券/商品风格因子性价比评分（估值分/动量分/综合信号等）',
        'source': '中证指数 + 东财行情，经性价比模型计算',
        'update': '每日通过 GitHub Actions CI 自动更新（北京时间 21:30）',
        'scoring': False,
    },
    'factor_scores_test': {
        'name': '风格因子评分测试表',
        'desc': 'factor_scores 的测试副本，结构一致，抓取数据先写入此表验证',
        'source': 'CI 抓取流程写入（验证用）',
        'update': '每次抓取数据时先写入此表',
        'scoring': False,
    },
    'style_factors': {
        'name': '风格因子明细表',
        'desc': '风格因子原始明细数据（指数代码/名称/PE/PB/历史分位/收益等多维度）',
        'source': '中证指数官网 + 蛋卷/东财估值',
        'update': '每日通过 GitHub Actions CI 自动更新（北京时间 21:30）',
        'scoring': False,
    },
    'jqr_indicators': {
        'name': '特色指标表（生产）',
        'desc': '市场情绪特色指标（恐惧贪婪/估值温度计/新发基金/股债差/破净率/证券化率等）',
        'source': '东财 push2 + 蛋卷等公开数据计算',
        'update': '每日通过 GitHub Actions CI 自动更新（北京时间 21:30）',
        'scoring': False,
    },
    'jqr_indicators_test': {
        'name': '特色指标测试表',
        'desc': 'jqr_indicators 的测试副本，结构一致，抓取数据先写入此表验证',
        'source': 'CI 抓取流程写入（验证用）',
        'update': '每次抓取数据时先写入此表',
        'scoring': False,
    },
    'etf_returns': {
        'name': 'ETF 收益率表',
        'desc': 'ETF 各周期收益率数据（代码/名称/近1周~成立以来/规模等）',
        'source': '天天基金 / 东财行情',
        'update': '每日通过 GitHub Actions CI 自动更新（北京时间 21:30）',
        'scoring': False,
    },
    'fund_category_indices': {
        'name': '基金分类指数表',
        'desc': '各基金分类对应的指数行情（分类名/指数代码/点位/各周期收益）',
        'source': '腾讯行情 qt.gtimg.cn + 东财',
        'update': '每日通过 GitHub Actions CI 自动更新（北京时间 21:30）',
        'scoring': False,
    },
    'fund_scores_staging': {
        'name': '评分暂存表（staging）',
        'desc': 'fund_scores 的 staging 暂存表，每日抓取先写入此表，经 promote_staging.py 严格校验后原子切换到生产，通常为临时状态',
        'source': 'CI 抓取流程写入（staging 管道）',
        'update': '每日抓取时写入，promote 后清空',
        'scoring': True,
    },
}

# 无元数据表的通用兜底说明（自动发现的新表）
GENERIC_META = {
    'name': '数据表',
    'desc': '大厨先生 数据库表（暂无详细说明）',
    'source': '大厨先生',
    'update': '按需更新',
    'scoring': False,
}

# 敏感表 / 含用户数据 — 仅在"我的"页面已登录时可见下载
SENSITIVE_TABLES = {'user_portfolios', 'user_profiles'}

import time

# ===== 导出 =====
def get_table_data(table_name):
    """分页获取表数据，带重试和超时处理"""
    all_rows = []
    offset = 0
    limit = 1000  # Supabase REST API default max rows per request
    max_retries = 5

    while True:
        success = False
        for attempt in range(max_retries):
            try:
                resp = requests.get(
                    f'{SUPABASE_URL}/rest/v1/{table_name}?select=*&limit={limit}&offset={offset}',
                    headers=HEADERS, timeout=120
                )
                if resp.status_code == 200:
                    success = True
                    break
                else:
                    print(f'  ⚠️ {table_name}: HTTP {resp.status_code} (attempt {attempt+1}/{max_retries})')
            except Exception as e:
                print(f'  ⚠️ {table_name}: {type(e).__name__} (attempt {attempt+1}/{max_retries})')
            if attempt < max_retries - 1:
                time.sleep(3 * (attempt + 1))
        
        if not success:
            print(f'  ❌ {table_name}: Failed after {max_retries} retries at offset {offset}')
            break
        
        rows = resp.json()
        if not rows:
            break
        all_rows.extend(rows)
        offset += len(rows)
        print(f'  📥 {table_name}: {offset} rows...')
        if len(rows) < limit:
            break
        time.sleep(0.3)  # Small delay to avoid rate limiting
    
    return all_rows

# ===== 列顺序定义（按用户要求的展示顺序） =====
FUND_SCORES_COL_ORDER = [
    'c','n','fund_manager','company','t0','t1','fund_scale','manage_fee',
    'ytd','r0w','r1m','r3m','r1y','r3y','r5y','r7y','r10y','return_all',
    'dd1y','dd2y','dd3y','dd5y',
    'sr1y','sr2y','sr3y','sr5y',
    'k0w','k1m','k3m','k6m','k1','k2','k3','k5','k_all','score_grade',
]
COLUMN_ORDER = {
    'fund_scores': FUND_SCORES_COL_ORDER,
    'fund_scores_test': FUND_SCORES_COL_ORDER,
}

# 列名中英文映射（Excel 表头用中文）
COLUMN_NAMES = {
    'c': '基金代码',
    'n': '基金名称',
    'fund_manager': '基金经理',
    'company': '管理人',
    't0': '一级分类',
    't1': '二级分类',
    't1_tt': '天天分类',
    'fund_scale': '基金规模(亿元)',
    'manage_fee': '管理费/y',
    'risk_level': '风险等级',
    'ytd': '今年以来',
    'r0w': '近1周',
    'r1m': '近1月',
    'r3m': '近3月',
    'r6m': '近6月',
    'r1y': '近1年',
    'r2y': '近2年',
    'r3y': '近3年',
    'r5y': '近5年',
    'r7y': '近7年',
    'r10y': '近10年',
    'return_all': '成立以来',
    'dd1y': '最大回撤1y',
    'dd2y': '最大回撤2y',
    'dd3y': '最大回撤3y',
    'dd5y': '最大回撤5y',
    'sr1y': '夏普比率1y',
    'sr2y': '夏普比率2y',
    'sr3y': '夏普比率3y',
    'sr5y': '夏普比率5y',
    'k0w': '评分_近1周',
    'k1m': '评分_近1月',
    'k3m': '评分_近3月',
    'k6m': '评分_近6月',
    'k1': '评分_近1年',
    'k2': '评分_近2年',
    'k3': '评分_近3年',
    'k5': '评分_近5年',
    'k7': '评分_近7年',
    'k10': '评分_近10年',
    'k_all': '综合评分',
    'score_grade': '评级',
    'daily_change': '日涨跌',
    'sg': '申购状态',
    'holders_count': '持有人数',
    'total_manage_scale': '经理总规模',
    'id': 'ID',
}

def export_to_excel(table_name, rows, output_path):
    """导出为 Excel，含数据说明 sheet"""
    from openpyxl.styles import Font, Alignment, Border, Side
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = table_name[:31]
    
    if not rows:
        ws.append(['(空表)'])
    else:
        # 列名：优先使用定义的顺序，否则用自然顺序
        col_order = COLUMN_ORDER.get(table_name)
        if col_order:
            # 只保留实际存在的列
            columns = [c for c in col_order if c in rows[0]]
            # 追加未在顺序中定义的新列
            extra = [c for c in rows[0].keys() if c not in columns]
            columns.extend(extra)
        else:
            columns = list(rows[0].keys())
        # 使用中文表头（如有映射），否则用原始列名
        headers = [COLUMN_NAMES.get(c, c) for c in columns]
        ws.append(headers)
        
        # 加粗表头
        for col_idx in range(1, len(columns) + 1):
            ws.cell(row=1, column=col_idx).font = Font(bold=True)
        
        # 数据行
        import json as _json
        def safe_val(v):
            if v is None: return ''
            if isinstance(v, (int, float, str, bool)): return v
            try: return _json.dumps(v, ensure_ascii=False)
            except: return str(v)
        
        for row in rows:
            ws.append([safe_val(row.get(col, '')) for col in columns])
        
        # 冻结首行
        ws.freeze_panes = 'A2'
    
    # ===== 添加"数据说明" sheet =====
    ws_meta = wb.create_sheet('数据说明')
    meta = TABLES.get(table_name) or dict(GENERIC_META, name=table_name)
    export_time = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    
    # 样式
    header_font = Font(name='微软雅黑', bold=True, size=14, color='1d70b8')
    label_font = Font(name='微软雅黑', bold=True, size=11)
    value_font = Font(name='微软雅黑', size=11)
    note_font = Font(name='微软雅黑', size=10, color='666666')
    wrap_align = Alignment(wrap_text=True, vertical='top')
    
    ws_meta.column_dimensions['A'].width = 18
    ws_meta.column_dimensions['B'].width = 80
    
    row_idx = 1
    
    # 标题
    ws_meta.cell(row=row_idx, column=1, value='大厨先生 数据说明').font = header_font
    ws_meta.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=2)
    row_idx += 2
    
    # 基本信息
    info_rows = [
        ('表名', table_name),
        ('中文名称', meta.get('name', '')),
        ('说明', meta.get('desc', '')),
        ('数据来源', meta.get('source', '')),
        ('更新频率', meta.get('update', '')),
        ('导出时间', export_time),
        ('行数', len(rows)),
    ]
    for label, value in info_rows:
        ws_meta.cell(row=row_idx, column=1, value=label).font = label_font
        c = ws_meta.cell(row=row_idx, column=2, value=str(value))
        c.font = value_font
        c.alignment = wrap_align
        row_idx += 1
    
    # 评分表增加评分说明
    if meta.get('scoring'):
        row_idx += 1
        ws_meta.cell(row=row_idx, column=1, value='评分方法 (V7)').font = Font(name='微软雅黑', bold=True, size=13, color='1d70b8')
        ws_meta.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=2)
        row_idx += 1
        
        scoring_notes = [
            ('算法版本', 'V7 — 收益 50% + 回撤 25% + 夏普 25%'),
            ('数据来源', 'FundGuideapi（阶段收益率 r0w~r5y）+ pingzhongdata（回撤 dd1y~dd5y、夏普 sr1y~sr5y）+ rankhandler（货币基金）'),
            ('百分位排名', '全市场基金按各指标降序排名，percentile = (1 - rank/(N-1)) × 100，范围 0~100'),
            ('短周期 k0w/k1m/k3m/k6m', '仅用收益率百分位排名：k_short = ret_percentile'),
            ('长周期 k1/k2/k3/k5', '三维度加权：k_long = 50% × ret_percentile + 25% × dd_percentile + 25% × sr_percentile'),
            ('综合评分 k_all', 'k_all = (k0w×5 + k1m×5 + k3m×10 + k6m×15 + k1×20 + k2×20 + k3×15 + k5×10) / total_weight（仅有效周期参与）'),
            ('评级 score_grade', '按 k_all 百分位分级：green(前20%) > blue(20%-50%) > orange(后50%) > gray(无数据)'),
            ('回撤计算', 'dd_max = -max((peak - nav[i]) / peak) × 100，负数百分比（如 -15.23 表示最大回撤 15.23%）'),
            ('夏普计算', 'Sharpe = (E[Rdaily] - Rf) / σdaily × √250，无风险利率 Rf = 2%/年 = 0.02/250 = 0.00008'),
            ('周期权重', 'k0w:5%, k1m:5%, k3m:10%, k6m:15%, k1:20%, k2:20%, k3:15%, k5:10%（总和=100，天然归一化）'),
        ]
        
        for label, value in scoring_notes:
            ws_meta.cell(row=row_idx, column=1, value=label).font = label_font
            c = ws_meta.cell(row=row_idx, column=2, value=str(value))
            c.font = value_font
            c.alignment = wrap_align
            row_idx += 1
    
    # 免责声明
    row_idx += 1
    ws_meta.cell(row=row_idx, column=1, value='免责声明').font = Font(name='微软雅黑', bold=True, size=11, color='999999')
    row_idx += 1
    disclaimer = (
        '本数据由 大厨先生 通过公开数据接口自动采集和计算，仅供参考，不构成任何投资建议。'
        '数据可能存在延迟或误差，请以天天基金等官方平台实时数据为准。'
        '投资有风险，入市需谨慎。'
    )
    c = ws_meta.cell(row=row_idx, column=1, value=disclaimer)
    c.font = note_font
    c.alignment = Alignment(wrap_text=True)
    ws_meta.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=2)
    
    wb.save(output_path)
    return len(rows)

def discover_tables():
    """通过 Supabase Management API 自动发现 public schema 下所有基础表。
    成功则返回表名列表；无 PAT 或失败则返回 None（回退到 TABLES 硬编码清单）。
    这样后续新建的表会被自动纳入数据下载中心，无需改代码。"""
    pat = os.environ.get('SUPABASE_MGMT_TOKEN') or os.environ.get('SUPABASE_PAT')
    if not pat:
        print('  ℹ️ 未提供 SUPABASE_MGMT_TOKEN/PAT，跳过自动发现，使用内置表清单')
        return None
    ref = SUPABASE_URL.replace('https://', '').replace('http://', '').split('.')[0]
    sql = ("select table_name from information_schema.tables "
           "where table_schema='public' and table_type='BASE TABLE' order by table_name;")
    try:
        r = requests.post(
            f'https://api.supabase.com/v1/projects/{ref}/database/query',
            headers={'Authorization': f'Bearer {pat}', 'Content-Type': 'application/json'},
            json={'query': sql}, timeout=60,
        )
        if r.status_code == 200:
            names = [row['table_name'] for row in r.json()]
            print(f'  🔎 自动发现 {len(names)} 张 public 表')
            return names
        print(f'  ⚠️ 自动发现失败 HTTP {r.status_code}，回退内置表清单')
    except Exception as e:
        print(f'  ⚠️ 自动发现异常 {type(e).__name__}，回退内置表清单')
    return None


def main():
    print(f'📊 导出 dachu 数据库全部表到 {OUTPUT_DIR}/')
    print(f'⏰ {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}\n')
    
    # 优先自动发现全部 public 表（含后续新建表）；失败则回退内置清单
    discovered = discover_tables()
    table_list = sorted(set(discovered) | set(TABLES.keys())) if discovered else sorted(TABLES.keys())
    print(f'📋 待导出 {len(table_list)} 张表\n')
    
    results = {}
    total_size = 0
    
    for table_name in table_list:
        print(f'⬇️ 导出 {table_name} ...')
        rows = get_table_data(table_name)
        
        output_path = os.path.join(OUTPUT_DIR, f'{table_name}.xlsx')
        count = export_to_excel(table_name, rows, output_path)
        
        size_mb = os.path.getsize(output_path) / (1024 * 1024)
        total_size += size_mb
        meta = TABLES.get(table_name) or dict(GENERIC_META, name=table_name)
        results[table_name] = {
            'rows': count,
            'size_mb': size_mb,
            'path': f'/downloads/{table_name}.xlsx',
            'name': meta.get('name', table_name),
            'desc': meta.get('desc', ''),
            'sensitive': table_name in SENSITIVE_TABLES,
        }
        print(f'  ✅ {table_name}: {count} rows, {size_mb:.2f}MB → {table_name}.xlsx')
    
    # 保存 JSON 索引文件（供前端读取）
    index_path = os.path.join(OUTPUT_DIR, 'index.json')
    with open(index_path, 'w') as f:
        json.dump({
            'updated_at': datetime.now().isoformat(),
            'tables': results,
        }, f, ensure_ascii=False, indent=2)
    
    print(f'\n✅ 全部完成！{len(results)} 张表，总大小 {total_size:.2f}MB')
    print(f'📋 索引文件: {index_path}')

if __name__ == '__main__':
    main()
